#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""運営ログ（非公開）→ ops.json（公開）変換

Google スプレッドシート「PG運営ログ（非公開）」の内容を、公開してよい部分だけに絞って
ops.json に落とす。**公開判断のロジックはこのファイル1箇所だけに置く。**
（fetch_ops.py も同じ関数を import して使う。二重実装にしないこと）

シート構成:
  宿題       内容 / 担当 / 期限 / 状態 / 公開 / 備考
  決定事項   内容 / 公開 / 備考
  施策       開始日 / 内容 / 公開 / 備考
  担当マスタ 内部名 / 公開表記 / 備考
  議事録     会議名 / 備考（URL・注意点）        ← ops.json には出さない

このファイルを直接実行する場合（手動運用）:
    python3 build_ops_json.py <入力ファイル> [出力先=ops.json]

    入力は次の2形式を自動判別する。
      1. Markdown のパイプ表が複数連なったもの … Drive コネクタの read_file_content の出力
      2. .xlsx … スプレッドシートをダウンロードした場合（openpyxl が必要）

自動運用（3時間おき）は fetch_ops.py が Sheets API から読んで同じ関数を呼ぶ。

設計の要点:
  * 公開列が "○" の行だけを出力する。空欄・× は出力しない（**既定は非公開**）。
  * 必須の列名が欠けていたら例外にする。黙って空の ops.json を作らない。
  * 担当者の実名は 担当マスタ に従って公開表記へ置き換える。担当列だけでなく本文中の実名も置換する。
  * マスタに無い「〜さん / 〜様 / 〜氏」が本文に残っていたら「担当者」に潰し、警告を出す（社外名の流出防止）。
  * 備考は出力しない（機微が混ざりやすいため）。
  * 非公開にした件数だけは出力する（「見えていない項目がある」ことを隠さないため）。
"""
import json
import re
import sys
from datetime import datetime, timedelta, timezone

JST = timezone(timedelta(hours=9))

# 公開とみなす表記（全角/半角・表記ゆれを吸収する。ここに無いものは全部「非公開」）
PUBLIC_TOKENS = {"○", "◯", "〇", "◎", "o", "yes", "y", "true", "1", "公開", "可"}

SEP_RE = re.compile(r"^:?-+:?$")  # Drive の出力は `:-:`（ハイフン1本）
HONORIFIC = re.compile(r"[぀-ヿ一-鿿A-Za-zＡ-Ｚａ-ｚ]{1,8}(?:さん|様|氏)")

# シートごとの必須列。1つでも欠けたらエラーにする（列名を変えたら気づけるように）
REQUIRED = {
    "宿題":       ["内容", "担当", "期限", "状態", "公開"],
    "決定事項":   ["内容", "公開"],
    "施策":       ["開始日", "内容", "公開"],
    "担当マスタ": ["内部名", "公開表記"],
}
STATUS_ORDER = {"進行中": 0, "未着手": 1, "保留": 2, "完了": 3, "見送り": 4}


class OpsError(Exception):
    """ops.json を作れない致命的な問題。呼び出し側は既存ファイルを上書きしないこと。"""


def norm(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return re.sub(r"\\([_*`\[\]|])", r"\1", s)  # Markdown のエスケープを戻す


# ──────────────────────────────────────────────────────────
# 公開用データの組み立て（ここが唯一の公開判断ロジック）
# ──────────────────────────────────────────────────────────
def build_ops(sheets):
    """sheets = {"宿題":[{列名:値,...},...], "決定事項":[...], "施策":[...], "担当マスタ":[...]}
    戻り値: (ops_dict, warnings)"""
    missing = [k for k in REQUIRED if k not in sheets]
    if missing:
        raise OpsError("シートが見つかりません: " + " / ".join(missing))

    for name, cols in REQUIRED.items():
        rows = sheets[name]
        if not rows:
            continue
        have = set(rows[0].keys())
        lack = [c for c in cols if c not in have]
        if lack:
            raise OpsError(f"[{name}] 必須の列がありません: {' / '.join(lack)}"
                           f"（実際の列: {' / '.join(sorted(have))}）")

    warnings = []

    owner_map = {}
    for r in sheets["担当マスタ"]:
        if r.get("内部名"):
            owner_map[r["内部名"]] = r.get("公開表記") or r["内部名"]
    if not owner_map:
        raise OpsError("担当マスタが空です。実名を公開表記に置き換えられないため中止します")

    def mask(text):
        """本文中の実名を公開表記に置換し、未知の人名は潰す"""
        if not text:
            return text
        for internal, public in owner_map.items():
            if internal and internal in text:
                text = text.replace(internal, public)

        def _sub(m):
            warnings.append(f"担当マスタに無い人名 '{m.group(0)}' を本文で検出 → '担当者' に置換しました")
            return "担当者"

        return HONORIFIC.sub(_sub, text)

    def pub_owner(name):
        if not name:
            return "未定"
        if name in owner_map:
            return owner_map[name]
        warnings.append(f"担当マスタに '{name}' がありません → '担当者' として出力しました")
        return "担当者"

    def is_public(row):
        return row.get("公開", "").strip().lower() in {t.lower() for t in PUBLIC_TOKENS}

    hidden = {"action_items": 0, "decisions": 0, "measures": 0}

    action_items = []
    for r in sheets["宿題"]:
        if not r.get("内容"):
            continue
        if not is_public(r):
            hidden["action_items"] += 1
            continue
        action_items.append({
            "text":   mask(r["内容"]),
            "owner":  pub_owner(r.get("担当", "")),
            "due":    r.get("期限", ""),
            "status": r.get("状態") or "未着手",
        })

    decisions = []
    for r in sheets["決定事項"]:
        if not r.get("内容"):
            continue
        if not is_public(r):
            hidden["decisions"] += 1
            continue
        decisions.append({"text": mask(r["内容"])})

    measures = []
    for r in sheets["施策"]:
        if not r.get("内容"):
            continue
        if not is_public(r):
            hidden["measures"] += 1
            continue
        measures.append({"start": r.get("開始日", ""), "text": mask(r["内容"])})

    action_items.sort(key=lambda a: (STATUS_ORDER.get(a["status"], 9), a["due"]))

    ops = {
        "updated_at": datetime.now(JST).strftime("%Y-%m-%d"),
        "source": "PG運営ログ（非公開・Google スプレッドシート）",
        "note": "公開列が○の項目のみ。担当は役割表記に置き換え済み。",
        "hidden_counts": hidden,
        "action_items": action_items,
        "decisions": decisions,
        "measures": measures,
    }
    return ops, warnings


def write_ops(ops, path="ops.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(ops, f, separators=(",", ":"), ensure_ascii=False)


def report(ops, warnings, path="ops.json"):
    h = ops["hidden_counts"]
    print(f"{path}: 宿題{len(ops['action_items'])}件 / 決定事項{len(ops['decisions'])}件 / 施策{len(ops['measures'])}件")
    print(f"  非公開のため除外: 宿題{h['action_items']} / 決定事項{h['decisions']} / 施策{h['measures']}")
    no_start = sum(1 for m in ops["measures"] if not m["start"])
    if no_start:
        print(f"  ※ 開始日が空の施策が {no_start} 件あります（数字の前後比較ができません）")
    for w in dict.fromkeys(warnings):
        print("  ⚠ " + w)


# ──────────────────────────────────────────────────────────
# 手動運用のための入力パース（Markdown表 / xlsx）
# ──────────────────────────────────────────────────────────
def _tables_to_sheets(tables):
    """表の集まりを、ヘッダの列名でシート名に振り分ける（並び順に依存しない）"""
    sheets = {}
    for t in tables:
        head, body = None, []
        for row in t:
            if head is None:
                if any(row):
                    head = row
                continue
            body.append({h: (row[i] if i < len(row) else "") for i, h in enumerate(head) if h})
        if not head:
            continue
        hs = set(head)
        if   "内部名" in hs:                        key = "担当マスタ"
        elif "会議名" in hs:                        key = "議事録"
        elif "開始日" in hs:                        key = "施策"
        elif "状態" in hs and "担当" in hs:          key = "宿題"
        elif "内容" in hs and "公開" in hs:          key = "決定事項"
        else:                                       continue
        sheets.setdefault(key, []).extend(body)
    return sheets


def parse_markdown_tables(text):
    tables, cur = [], []
    for line in text.splitlines():
        s = line.strip()
        if not s.startswith("|"):
            if cur:
                tables.append(cur); cur = []
            continue
        cells = [norm(c) for c in s.strip("|").split("|")]
        if cells and all(SEP_RE.fullmatch(c) for c in cells if c):
            continue
        if any(cells):
            cur.append(cells)
    if cur:
        tables.append(cur)
    return tables


def parse_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise OpsError("xlsx を読むには openpyxl が必要です: pip install openpyxl --break-system-packages")
    wb = load_workbook(path, data_only=True)
    out = []
    for ws in wb:
        rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(r)]
        if rows:
            out.append(rows)
    return out


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src, dst = sys.argv[1], (sys.argv[2] if len(sys.argv) > 2 else "ops.json")
    try:
        if src.lower().endswith((".xlsx", ".xlsm")):
            tables = parse_xlsx(src)
        else:
            with open(src, "r", encoding="utf-8") as f:
                tables = parse_markdown_tables(f.read())
        ops, warnings = build_ops(_tables_to_sheets(tables))
    except OpsError as e:
        sys.exit(f"[ERROR] {e}\n（ops.json は書き換えていません）")
    write_ops(ops, dst)
    report(ops, warnings, dst)
    if warnings:
        print("\n⚠ 警告が出ています。担当マスタの更新漏れか、社外名の混入です。止まって報告してください。")


if __name__ == "__main__":
    main()
